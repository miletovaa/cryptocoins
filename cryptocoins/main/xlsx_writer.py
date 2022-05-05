import time
from printInfo import get_prices
import pickle
import openpyxl
import subprocess
from getY1 import send_message_Y1


def save_xlsx_file():
    subprocess.Popen(['libreoffice', '--calc', './result.xlsx'])
    time.sleep(5)
    subprocess.Popen(['xdotool', 'key', 'ctrl+s'])
    time.sleep(0.5)
    subprocess.Popen(['xdotool', 'key', 'Return'])
    time.sleep(0.5)
    subprocess.Popen(['xdotool', 'key', 'ctrl+F4'])
    time.sleep(0.5)
    subprocess.Popen(['xdotool', 'key', 'ctrl+F4'])
    time.sleep(0.5)
    subprocess.Popen(['xdotool', 'key', 'ctrl+F4'])
    print('[Saved]')


def write():
    workbook = openpyxl.load_workbook('result.xlsx')
    worksheet = workbook.active

    for i in range(3, 43):
        for j in range(1, 57):
            try:
                worksheet.cell(i, j).value = ''
            except:
                continue

    for i in range(3, 43):
        for j in range(105, 125):
            try:
                worksheet.cell(i, j).value = ''
            except:
                continue

    data = get_prices()

    all_coins = []
    for el in data:
        if data[el] != None:
            for el2 in data[el]:
                all_coins.append(el2[0])
    all_coins = list(set(all_coins))

    for el in all_coins:
        worksheet.cell(all_coins.index(el) + 3, 1).value = el.upper()

    configs = pickle.load(open('configs.pickle', 'rb'))
    for i in range(len(configs['values'])):
        worksheet.cell(row=3, column=i * 11 + 2).value = configs['values'][i]

    for i in range(3, 43):
        try:
            for j in range(len(configs['values'])):
                worksheet.unmerge_cells(start_row=3, start_column=j * 11 + 2, end_row=i, end_column=j * 11 + 2)
            break
        except:
            continue

    for j in range(len(configs['values'])):
        worksheet.merge_cells(start_row=3, start_column=j * 11 + 2, end_row=len(all_coins) + 2, end_column=j * 11 + 2)

    if data['junoswap']:
        for el in data['junoswap']:
            for i in range(1, len(el)):
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 6).value = el[i][1] if el[i][1] != 0 else ''
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 11).value = el[i][2] if el[i][2] != 0 else ''
    if data['sifchain']:
        for el in data['sifchain']:
            for i in range(1, len(el)):
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 5).value = el[i][1] if el[i][1] != 0 else ''
                # worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 10).value = el[i][2] if el[i][2] != 0 else ''
                worksheet.cell(all_coins.index(el[0]) + 3, 105 + 2 * (i - 1)).value = el[i][3] if el[i][3] != 0 else ''
                worksheet.cell(all_coins.index(el[0]) + 3, 106 + 2 * (i - 1)).value = el[i][4] if el[i][4] != 0 else ''
    if data['marbledao']:
        for el in data['marbledao']:
            for i in range(1, len(el)):
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 7).value = el[i][1] if el[i][1] != 0 else ''
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 12).value = el[i][2] if el[i][2] != 0 else ''
    if data['crescent']:
        for el in data['crescent']:
            for i in range(1, len(el)):
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 4).value = el[i][1] if el[i][1] != 0 else ''
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 9).value = el[i][2] if el[i][2] != 0 else ''
    if data['osmosis']:
        for el in data['osmosis']:
            for i in range(1, len(el)):
                worksheet.cell(all_coins.index(el[0]) + 3, 11 * (i - 1) + 3).value = el[i][1] if el[i][1] != 0 else ''

    workbook.save('result.xlsx')
    workbook.close()

    save_xlsx_file()

    send_message_Y1()
