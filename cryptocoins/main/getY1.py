import subprocess
import openpyxl
from multiprocessing import Process
from swapCoins import SwapCoins
from send_message import send_message


def send_message_Y1():
    wb = openpyxl.load_workbook('result.xlsx', data_only=True)
    sheet = wb.active
    wb.iso_dates = True
    data = sheet['Y1'].value

    print(f'[Y1 - {data}]')

    if data is not None:
        i = 1
        send_message(data)
        while True:
            if sheet.cell(2, 78 + 2 * (i - 1) + 1).value is not None:
                func1 = None
                func2 = None
                if sheet.cell(2, 78 + 2 * (i - 1) + 1).value == 'Osmosis':
                    func1 = Process(target=swapOsmosis, args=(sheet.cell(3, 78 + 2 * (i - 1) + 1).value.lower(),
                                                              sheet.cell(5, 78 + 2 * (i - 1) + 1).value,
                                                              sheet.cell(4, 78 + 2 * (i - 1) + 1).value.lower()))
                elif sheet.cell(2, 78 + 2 * (i - 1) + 1).value == 'Sifchain':
                    func1 = Process(target=swapSifchain, args=(sheet.cell(3, 78 + 2 * (i - 1) + 1).value.lower(),
                                                               sheet.cell(5, 78 + 2 * (i - 1) + 1).value,
                                                               sheet.cell(4, 78 + 2 * (i - 1) + 1).value.lower()))
                elif sheet.cell(2, 78 + 2 * (i - 1) + 1).value == 'Junoswap':
                    func1 = Process(target=swapJunoSwap, args=(sheet.cell(3, 78 + 2 * (i - 1) + 1).value.lower(),
                                                               sheet.cell(5, 78 + 2 * (i - 1) + 1).value,
                                                               sheet.cell(4, 78 + 2 * (i - 1) + 1).value.lower()))
                elif sheet.cell(2, 78 + 2 * (i - 1) + 1).value == 'Crescent':
                    func1 = Process(target=swapCrescent, args=(sheet.cell(3, 78 + 2 * (i - 1) + 1).value.lower(),
                                                               sheet.cell(5, 78 + 2 * (i - 1) + 1).value,
                                                               sheet.cell(4, 78 + 2 * (i - 1) + 1).value.lower()))
                if sheet.cell(2, 78 + 2 * (i - 1) + 2).value == 'Osmosis':
                    func2 = Process(target=swapOsmosis, args=(sheet.cell(3, 78 + 2 * (i - 1) + 2).value.lower(),
                                                              sheet.cell(5, 78 + 2 * (i - 1) + 2).value,
                                                              sheet.cell(4, 78 + 2 * (i - 1) + 2).value.lower()))
                elif sheet.cell(2, 78 + 2 * (i - 1) + 2).value == 'Sifchain':
                    func2 = Process(target=swapSifchain, args=(sheet.cell(3, 78 + 2 * (i - 1) + 2).value.lower(),
                                                               sheet.cell(5, 78 + 2 * (i - 1) + 2).value,
                                                               sheet.cell(4, 78 + 2 * (i - 1) + 2).value.lower()))
                elif sheet.cell(2, 78 + 2 * (i - 1) + 2).value == 'Junoswap':
                    func2 = Process(target=swapJunoSwap, args=(sheet.cell(3, 78 + 2 * (i - 1) + 2).value.lower(),
                                                               sheet.cell(5, 78 + 2 * (i - 1) + 2).value,
                                                               sheet.cell(4, 78 + 2 * (i - 1) + 2).value.lower()))
                elif sheet.cell(2, 78 + 2 * (i - 1) + 2).value == 'Crescent':
                    func2 = Process(target=swapCrescent, args=(sheet.cell(3, 78 + 2 * (i - 1) + 2).value.lower(),
                                                               sheet.cell(5, 78 + 2 * (i - 1) + 2).value,
                                                               sheet.cell(4, 78 + 2 * (i - 1) + 2).value.lower()))
                func1.start()
                func2.start()
                func1.join()
                func2.join()
                send_message('Transactions were completed successfully!')
                func1.close()
                func2.close()
                i += 1
            else:
                break


def swapOsmosis(main_coin, value, second_coin):
    # subprocess.Popen(
    #     ['python3', 'swapOsmosis.py', str(main_coin),
    #      str(value),
    #      str(second_coin)])
    osmosis = SwapCoins.Osmosis()
    osmosis.swapCoins(main_coin, value, second_coin)


def swapJunoSwap(main_coin, value, second_coin):
    subprocess.Popen(
        ['python3', 'swapJunoSwap.py', str(main_coin),
         str(value),
         str(second_coin)])


def swapSifchain(main_coin, value, second_coin):
    # subprocess.Popen(
    #     ['python3', 'swapSifchain.py', str(main_coin),
    #      str(value),
    #      str(second_coin)])
    sifchain = SwapCoins.Sifchain()
    sifchain.swapCoins(main_coin, value, second_coin)


def swapCrescent(main_coin, value, second_coin):
    subprocess.Popen(
        ['python3', 'swapCrescent.py', str(main_coin),
         str(value),
         str(second_coin)])
